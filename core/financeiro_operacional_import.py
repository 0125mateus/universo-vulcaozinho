"""Importação de planilhas XLSX — compras e atrações."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

import openpyxl

from .models import (
    ItemCompraSemanal,
    PagamentoAtracao,
    PeriodoOperacional,
    StatusPagamentoOperacional,
    TipoPeriodoOperacional,
)


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    return str(value).strip()


def _parse_decimal(value) -> Decimal:
    if value is None or value == '':
        return Decimal('0')
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace('R$', '').replace('.', '').replace(',', '.')
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal('0')


def _parse_int(value) -> int:
    try:
        return max(1, int(float(value)))
    except (TypeError, ValueError):
        return 1


def _normalize_link(value) -> str:
    text = _cell_str(value).lower()
    if text in ('', 'link'):
        return ''
    if text.startswith('http'):
        return text
    return ''


def importar_compras_xlsx(
    arquivo: BinaryIO,
    hotel,
    periodo: PeriodoOperacional | None = None,
    substituir: bool = False,
) -> tuple[int, list[str]]:
    """Importa planilha de compras semanais (materiais)."""
    erros: list[str] = []
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active
    if substituir and periodo:
        ItemCompraSemanal.objects.filter(hotel=hotel, periodo=periodo).delete()

    criados = 0
    ordem = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        descricao = _cell_str(row[0] if row else '')
        if not descricao or descricao.lower().startswith('valor total'):
            continue
        quantidade = _parse_int(row[1] if len(row) > 1 else 1)
        link = _normalize_link(row[2] if len(row) > 2 else '')
        if not link:
            raw_link = _cell_str(row[2] if len(row) > 2 else '')
            link = raw_link[:500] if raw_link and raw_link.lower() != 'link' else ''
        preco_unit = _parse_decimal(row[3] if len(row) > 3 else 0)
        if preco_unit <= 0 and len(row) > 7:
            preco_unit = _parse_decimal(row[7]) / quantidade if quantidade else Decimal('0')

        ordem += 1
        ItemCompraSemanal.objects.create(
            hotel=hotel,
            periodo=periodo,
            descricao=descricao[:300],
            quantidade=quantidade,
            link_fornecedor=link[:500],
            preco_unitario=preco_unit,
            ordem=ordem,
        )
        criados += 1

    wb.close()
    if criados == 0:
        erros.append('Nenhum item encontrado na planilha de compras.')
    return criados, erros


def importar_atracoes_xlsx(
    arquivo: BinaryIO,
    hotel,
    periodo: PeriodoOperacional | None = None,
    substituir: bool = False,
) -> tuple[int, list[str]]:
    """Importa planilha de atrações / pagamentos de artistas."""
    erros: list[str] = []
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        cells = [_cell_str(c).lower() for c in row[:8]]
        if 'artista' in cells and 'valor' in cells:
            header_row = i
            break

    if not header_row:
        wb.close()
        return 0, ['Cabeçalho não encontrado (esperado: Data, Artista, Atração, Valor, Chave PIX).']

    if substituir and periodo:
        PagamentoAtracao.objects.filter(hotel=hotel, periodo=periodo).delete()

    criados = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        data_label = _cell_str(row[0] if row else '')
        artista = _cell_str(row[1] if len(row) > 1 else '')
        atracao = _cell_str(row[2] if len(row) > 2 else '')
        valor = _parse_decimal(row[3] if len(row) > 3 else 0)
        chave_pix = _cell_str(row[4] if len(row) > 4 else '')

        if not artista:
            continue
        if valor <= 0 and not atracao:
            continue

        data_evento = None
        if isinstance(row[0], datetime):
            data_evento = row[0].date()
        elif re.search(r'\d{2}/\d{2}', data_label):
            try:
                data_evento = datetime.strptime(data_label[:10], '%d/%m/%Y').date()
            except ValueError:
                pass

        PagamentoAtracao.objects.create(
            hotel=hotel,
            periodo=periodo,
            data_label=data_label[:80],
            data_evento=data_evento,
            artista=artista[:200],
            atracao=atracao[:200],
            valor=valor,
            chave_pix=chave_pix[:200],
            status=StatusPagamentoOperacional.PENDENTE,
        )
        criados += 1

    wb.close()
    if criados == 0:
        erros.append('Nenhum pagamento de atração encontrado na planilha.')
    return criados, erros


def importar_eventos_recreacao_xlsx(
    arquivo: BinaryIO,
    hotel,
    periodo: PeriodoOperacional | None = None,
    substituir: bool = False,
) -> tuple[int, list[str]]:
    """Importa aba mensal da planilha de valores recreação (eventos)."""
    erros: list[str] = []
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        cells = [_cell_str(c).lower() for c in row[:14]]
        if 'prestador' in ' '.join(cells) and 'orçamento' in ' '.join(cells) or 'orcamento' in ' '.join(cells):
            header_row = i
            break

    if not header_row:
        wb.close()
        return 0, ['Cabeçalho de eventos não encontrado na aba.']

    if substituir and periodo:
        PagamentoAtracao.objects.filter(hotel=hotel, periodo=periodo).delete()

    criados = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        pacote = _cell_str(row[0] if row else '')
        evento = _cell_str(row[2] if len(row) > 2 else '')
        tipo_servico = _cell_str(row[3] if len(row) > 3 else '')
        artista = _cell_str(row[4] if len(row) > 4 else '')
        responsavel = _cell_str(row[5] if len(row) > 5 else '')
        data_label = _cell_str(row[6] if len(row) > 6 else '')
        data_ini = row[7] if len(row) > 7 else None
        horario = _cell_str(row[9] if len(row) > 9 else '')
        local_dept = _cell_str(row[10] if len(row) > 10 else '')
        valor = _parse_decimal(row[11] if len(row) > 11 else 0)
        autorizacao = _cell_str(row[12] if len(row) > 12 else '')

        if not artista and not evento:
            continue

        data_evento = None
        if isinstance(data_ini, datetime):
            data_evento = data_ini.date()
        elif isinstance(data_ini, date):
            data_evento = data_ini

        status = StatusPagamentoOperacional.PENDENTE
        if autorizacao.upper() == 'AUTORIZADO':
            status = StatusPagamentoOperacional.AUTORIZADO

        PagamentoAtracao.objects.create(
            hotel=hotel,
            periodo=periodo,
            pacote=pacote[:120],
            evento=evento[:200],
            tipo_servico=tipo_servico[:120],
            artista=artista[:200] or evento[:200],
            atracao=tipo_servico[:200],
            responsavel=responsavel[:120],
            data_label=data_label[:80],
            data_evento=data_evento,
            horario=horario[:40],
            local_dept=local_dept[:120],
            valor=valor,
            autorizacao_diretoria=autorizacao[:80],
            status=status,
        )
        criados += 1

    wb.close()
    if criados == 0:
        erros.append('Nenhum evento encontrado na planilha.')
    return criados, erros
