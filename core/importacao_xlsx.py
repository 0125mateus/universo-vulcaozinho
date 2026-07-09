"""Importação de planilhas Excel — programação diária e eventos especiais."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

import openpyxl

from django.db import transaction
from django.utils import timezone

from .models import (
    Atividade,
    CategoriaProgramacao,
    EventoRecreacao,
    Hotel,
    LocalAtividade,
    ProgramacaoDiaria,
    Recreador,
    StatusEventoRecreacao,
)

# Abas ignoradas na planilha de eventos
ABAS_EVENTOS_IGNORAR = {'dados', 'página11', 'pagina11'}

# Mapeamento nome da aba → código da categoria (padrão rede)
MAPA_ABA_CATEGORIA = (
    (re.compile(r'adulto', re.I), 'adultos'),
    (re.compile(r'teen', re.I), 'boys-girls'),
    (re.compile(r'boys|girls|7.*12', re.I), 'vulcao-kids'),
    (re.compile(r'kids|4.*6', re.I), 'vulcao-kids'),
)

IGNORAR_ATIVIDADES = frozenset({
    'intervalo', 'saída', 'saida', 'intervalo para almoço', 'intervalo para almoco',
})

COL_ALIASES_PROG = {
    'data': ('data',),
    'dia': ('dia da semana', 'dia'),
    'hora': ('hora', 'horário', 'horario'),
    'atividade': ('atividade',),
    'local': ('local',),
    'musico': ('músico', 'musico'),
    'atividade_chuva': ('atividade (b) - chuva', 'atividade (b)- chuva', 'atividade b chuva'),
    'local_chuva': ('local (b) - chuva', 'local chuva'),
    'responsavel': ('responsável', 'responsavel'),
    'coordenador': ('coordenador',),
    'realizado': ('realizado',),
    'auditado': ('auditado por', 'auditado'),
}

COL_ALIASES_EVENTO = {
    'pacote': ('pacote',),
    'mes': ('mês', 'mes'),
    'orcamento': ('orçamento', 'orcamento'),
    'evento': ('evento',),
    'descricao': ('descrição', 'descricao'),
    'tipo': ('tipo serviço / fornecedor', 'tipo servico / fornecedor', 'tipo serviço'),
    'nivel': ('nível atração', 'nivel atracao', 'nível atração'),
    'contrato': ('contrato assinado?', 'contrato assinado'),
    'prestador': ('prestador / fornecedor', 'prestador'),
    'responsavel': ('responsável', 'responsavel'),
    'dia': ('dia',),
    'data_inicio': ('data inicial', 'data inicio'),
    'data_fim': ('data final',),
}


@dataclass
class ResultadoImportacao:
    criados: int = 0
    atualizados: int = 0
    ignorados: int = 0
    erros: list[str] = field(default_factory=list)
    abas_processadas: list[str] = field(default_factory=list)

    @property
    def total(self) -> int:
        return self.criados + self.atualizados

    def resumo(self) -> str:
        return (
            f'{self.criados} criados, {self.atualizados} atualizados, '
            f'{self.ignorados} ignorados, {len(self.erros)} erros'
        )


@dataclass
class LinhaProgramacao:
    """Uma linha lida da planilha de programação (sem gravar no banco)."""
    aba: str
    faixa_label: str
    categoria_codigo: str | None
    data: date | None
    hora: time | None
    atividade: str
    local: str = ''


def _norm_header(cell) -> str:
    if cell is None:
        return ''
    return re.sub(r'\s+', ' ', str(cell).strip().lower())


def _mapear_colunas(header_row, aliases: dict) -> dict[str, int]:
    headers = [_norm_header(c) for c in header_row]
    out = {}
    for key, names in aliases.items():
        for i, h in enumerate(headers):
            if h in names or any(n in h for n in names if len(n) > 4):
                out[key] = i
                break
    return out


def _cell(row, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_date(val) -> date | None:
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(val) -> time | None:
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.time().replace(second=0, microsecond=0)
    if isinstance(val, time):
        return val.replace(second=0, microsecond=0)
    s = str(val).strip()
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m:
        return time(int(m.group(1)), int(m.group(2)))
    return None


def _parse_decimal(val) -> Decimal | None:
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _extrair_local(texto: str) -> str:
    if not texto:
        return 'A definir'
    t = str(texto).strip()
    if 'local:' in t.lower():
        return t.split(':', 1)[-1].strip()[:120] or 'A definir'
    return t[:120]


def _categoria_por_aba(nome_aba: str) -> str | None:
    for pattern, codigo in MAPA_ABA_CATEGORIA:
        if pattern.search(nome_aba):
            return codigo
    return None


def faixa_label_da_aba(nome_aba: str) -> str:
    """Rótulo legível da faixa a partir do nome da aba Excel."""
    if re.search(r'melhor|idos', nome_aba, re.I):
        return 'Melhor Idade 60+'
    if re.search(r'kids|4.*6', nome_aba, re.I):
        return 'Kids 4–6 anos'
    if re.search(r'boys|girls|7.*12', nome_aba, re.I):
        return 'Boys & Girls 7–12'
    if re.search(r'teen', nome_aba, re.I):
        return 'Teens 13–17'
    if re.search(r'adulto', nome_aba, re.I):
        return 'Adultos 18–59'
    if re.match(r'^\d{4}$', nome_aba.strip()):
        return f'Grade {nome_aba}'
    return nome_aba.strip()


def extrair_linhas_programacao(arquivo: BinaryIO) -> tuple[list[LinhaProgramacao], list[str]]:
    """Lê planilha de programação sem importar — para análise e cruzamento."""
    linhas: list[LinhaProgramacao] = []
    erros: list[str] = []
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        if not _aba_eh_programacao(sheet_name):
            continue
        codigo_cat = _categoria_por_aba(sheet_name)
        faixa = faixa_label_da_aba(sheet_name)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        found = _find_header_row(rows, {'atividade', 'hora'})
        if not found:
            erros.append(f'Aba "{sheet_name}": cabeçalho não encontrado.')
            continue
        header_idx, cols = found
        for row in rows[header_idx + 1:]:
            nome_ativ = _cell(row, cols.get('atividade'))
            if not nome_ativ:
                continue
            nome_ativ = str(nome_ativ).strip()
            if nome_ativ.lower() in IGNORAR_ATIVIDADES:
                continue
            linhas.append(LinhaProgramacao(
                aba=sheet_name,
                faixa_label=faixa,
                categoria_codigo=codigo_cat,
                data=_parse_date(_cell(row, cols.get('data'))),
                hora=_parse_time(_cell(row, cols.get('hora'))),
                atividade=nome_ativ,
                local=_extrair_local(str(_cell(row, cols.get('local')) or '')),
            ))
    wb.close()
    return linhas, erros


def _aba_eh_programacao(nome_aba: str) -> bool:
    n = nome_aba.lower()
    if n in ('show',):
        return False
    if re.match(r'^\d{4}$', n):
        return True
    return _categoria_por_aba(nome_aba) is not None


def _get_or_create_local(hotel: Hotel, nome: str) -> LocalAtividade:
    nome = nome[:120] if nome else 'A definir'
    local, _ = LocalAtividade.objects.get_or_create(
        hotel=hotel,
        nome=nome,
        defaults={'capacidade_maxima': 50, 'ativo': True},
    )
    return local


def _get_or_create_atividade(
    hotel: Hotel,
    nome: str,
    categoria: CategoriaProgramacao | None,
    local: LocalAtividade,
) -> Atividade:
    atividade, _ = Atividade.objects.update_or_create(
        hotel=hotel,
        nome=nome[:120],
        defaults={
            'categoria': categoria,
            'local_padrao': local,
            'ativo': True,
            'duracao_minutos': 30,
        },
    )
    return atividade


def _parse_bool(val) -> bool:
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ('1', 'true', 'sim', 'x', 'ok', 'realizado')


def _find_header_row(rows: list, min_cols: set) -> tuple[int, dict[str, int]] | None:
    for i, row in enumerate(rows[:30]):
        cols = _mapear_colunas(row, COL_ALIASES_PROG if 'atividade' in min_cols else COL_ALIASES_EVENTO)
        aliases = COL_ALIASES_PROG if 'atividade' in min_cols else COL_ALIASES_EVENTO
        needed = min_cols
        if all(k in cols for k in needed):
            return i, cols
        # fallback: partial match for programação
        if 'atividade' in needed and 'atividade' in cols and 'hora' in cols:
            return i, cols
        if 'evento' in needed and 'evento' in cols and 'data_inicio' in cols:
            return i, cols
    return None


@transaction.atomic
def importar_programacao(hotel: Hotel, arquivo: BinaryIO, *, substituir_datas: bool = False) -> ResultadoImportacao:
    """Importa planilha de programação diária (padrão Nacional Inn)."""
    resultado = ResultadoImportacao()
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

    categorias = {c.codigo: c for c in CategoriaProgramacao.objects.all()}
    if not categorias:
        resultado.erros.append('Cadastre faixas etárias (seed_categorias) antes de importar.')
        return resultado

    recreador, _ = Recreador.objects.get_or_create(
        hotel=hotel,
        nome='Equipe Recreação',
        defaults={'ativo': True},
    )

    for sheet_name in wb.sheetnames:
        if not _aba_eh_programacao(sheet_name):
            continue

        codigo_cat = _categoria_por_aba(sheet_name)
        categoria = categorias.get(codigo_cat) if codigo_cat else categorias.get('adultos')

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        found = _find_header_row(rows, {'atividade', 'hora'})
        if not found:
            resultado.erros.append(f'Aba "{sheet_name}": cabeçalho não encontrado.')
            continue

        header_idx, cols = found
        resultado.abas_processadas.append(sheet_name)

        if substituir_datas:
            datas_aba = set()
            for row in rows[header_idx + 1:]:
                d = _parse_date(_cell(row, cols.get('data')))
                if d:
                    datas_aba.add(d)
            if datas_aba:
                ProgramacaoDiaria.objects.filter(hotel=hotel, data__in=datas_aba).delete()

        for row in rows[header_idx + 1:]:
            try:
                nome_ativ = _cell(row, cols.get('atividade'))
                if not nome_ativ:
                    resultado.ignorados += 1
                    continue
                nome_ativ = str(nome_ativ).strip()
                if nome_ativ.lower() in IGNORAR_ATIVIDADES:
                    resultado.ignorados += 1
                    continue

                data = _parse_date(_cell(row, cols.get('data')))
                hora_inicio = _parse_time(_cell(row, cols.get('hora')))
                if not data or not hora_inicio:
                    resultado.ignorados += 1
                    continue

                local_nome = _extrair_local(str(_cell(row, cols.get('local')) or ''))
                local = _get_or_create_local(hotel, local_nome)
                atividade = _get_or_create_atividade(hotel, nome_ativ, categoria, local)

                hora_fim = (
                    datetime.combine(date.today(), hora_inicio) + timedelta(minutes=30)
                ).time()

                chuva = _cell(row, cols.get('atividade_chuva'))
                local_chuva = _cell(row, cols.get('local_chuva'))
                resp = _cell(row, cols.get('responsavel'))
                coord = _cell(row, cols.get('coordenador'))
                musico = _cell(row, cols.get('musico'))
                realizado = _parse_bool(_cell(row, cols.get('realizado')))
                auditado = _cell(row, cols.get('auditado'))

                defaults = {
                    'hora_fim': hora_fim,
                    'atividade': atividade,
                    'categoria': categoria,
                    'recreador': recreador,
                    'vagas_total': local.capacidade_maxima,
                    'atividade_chuva': str(chuva).strip()[:200] if chuva else '',
                    'local_chuva_nome': _extrair_local(str(local_chuva or ''))[:200],
                    'responsavel_nome': str(resp).strip()[:120] if resp else '',
                    'coordenador_nome': str(coord).strip()[:120] if coord else '',
                    'musico_nome': str(musico).strip()[:120] if musico else '',
                    'realizado': realizado,
                    'auditado_por': str(auditado).strip()[:120] if auditado else '',
                }

                obj, created = ProgramacaoDiaria.objects.update_or_create(
                    hotel=hotel,
                    data=data,
                    hora_inicio=hora_inicio,
                    local=local,
                    defaults=defaults,
                )
                if created:
                    resultado.criados += 1
                else:
                    resultado.atualizados += 1
            except Exception as exc:
                resultado.erros.append(f'{sheet_name}: {exc}')

    wb.close()
    return resultado


@transaction.atomic
def importar_eventos(hotel: Hotel, arquivo: BinaryIO) -> ResultadoImportacao:
    """Importa planilha de eventos especiais (orçamento, fornecedor, status)."""
    resultado = ResultadoImportacao()
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in ABAS_EVENTOS_IGNORAR:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 7:
            continue

        found = _find_header_row(rows, {'evento', 'data_inicio'})
        if not found:
            continue

        header_idx, cols = found
        resultado.abas_processadas.append(sheet_name)

        for row in rows[header_idx + 1:]:
            try:
                nome = _cell(row, cols.get('evento'))
                if not nome:
                    resultado.ignorados += 1
                    continue
                nome = str(nome).strip()
                if nome.upper().startswith('TOTAL') or 'PLANILHA' in nome.upper():
                    resultado.ignorados += 1
                    continue

                data_inicio = _parse_date(_cell(row, cols.get('data_inicio')))
                if not data_inicio:
                    resultado.ignorados += 1
                    continue

                data_fim = _parse_date(_cell(row, cols.get('data_fim')))
                orcamento = _parse_decimal(_cell(row, cols.get('orcamento')))
                pacote = _cell(row, cols.get('pacote'))
                mes = _cell(row, cols.get('mes'))
                desc = _cell(row, cols.get('descricao'))
                tipo = _cell(row, cols.get('tipo'))
                nivel = _cell(row, cols.get('nivel'))
                contrato = _cell(row, cols.get('contrato'))
                prestador = _cell(row, cols.get('prestador'))
                resp = _cell(row, cols.get('responsavel'))
                dia = _cell(row, cols.get('dia'))

                defaults = {
                    'pacote': str(pacote).strip()[:80] if pacote else '',
                    'mes_referencia': str(mes).strip()[:40] if mes else '',
                    'descricao': str(desc).strip() if desc else '',
                    'tipo_servico': str(tipo).strip()[:80] if tipo else '',
                    'nivel_atracao': str(nivel).strip()[:60] if nivel else '',
                    'contrato_assinado': str(contrato).strip()[:40] if contrato else '',
                    'prestador': str(prestador).strip()[:120] if prestador else '',
                    'responsavel': str(resp).strip()[:120] if resp else '',
                    'dia_semana': str(dia).strip()[:20] if dia else '',
                    'data_fim': data_fim,
                    'orcamento': orcamento,
                    'status': StatusEventoRecreacao.AGENDADO,
                }

                obj, created = EventoRecreacao.objects.update_or_create(
                    hotel=hotel,
                    data_inicio=data_inicio,
                    nome=nome[:200],
                    defaults=defaults,
                )
                if created:
                    resultado.criados += 1
                else:
                    resultado.atualizados += 1
            except Exception as exc:
                resultado.erros.append(f'{sheet_name}: {exc}')

    wb.close()
    return resultado
