"""Exportação de ponto para RH — planilha Excel e PDF."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
from django.db.models import QuerySet
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Hotel, PontoBatida, TipoPontoBatida

_THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


@dataclass
class ResumoDiaRH:
    recreador_id: int
    recreador_nome: str
    telefone: str
    data: date
    primeira_entrada: str
    ultima_saida: str
    horas_trabalhadas: float
    qtd_entradas: int
    qtd_saidas: int
    qtd_extras: int
    observacao: str


def _fmt_hora(dt: datetime | None) -> str:
    if not dt:
        return '—'
    return timezone.localtime(dt).strftime('%H:%M')


def _horas_pares(batidas_ordenadas: list[PontoBatida]) -> tuple[float, str]:
    """Calcula horas a partir de pares entrada→saída. Retorna (horas, observação)."""
    total = 0.0
    aberta: PontoBatida | None = None
    obs: list[str] = []
    for b in batidas_ordenadas:
        if b.tipo == TipoPontoBatida.ENTRADA:
            if aberta:
                obs.append('Entrada duplicada sem saída')
            aberta = b
        elif b.tipo == TipoPontoBatida.SAIDA:
            if aberta:
                delta = b.registrado_em - aberta.registrado_em
                total += max(delta.total_seconds(), 0) / 3600.0
                aberta = None
            else:
                obs.append('Saída sem entrada')
    if aberta:
        obs.append('Entrada sem saída (turno aberto)')
    return round(total, 2), '; '.join(dict.fromkeys(obs)) if obs else 'OK'


def montar_resumos_rh(
    batidas: list[PontoBatida],
    *,
    data_inicio: date,
    data_fim: date,
) -> list[ResumoDiaRH]:
    """Um resumo por recreador × dia civil no intervalo."""
    por_chave: dict[tuple[int, date], list[PontoBatida]] = {}
    nomes: dict[int, tuple[str, str]] = {}
    for b in batidas:
        dia = timezone.localtime(b.registrado_em).date()
        if dia < data_inicio or dia > data_fim:
            continue
        key = (b.recreador_id, dia)
        por_chave.setdefault(key, []).append(b)
        nomes[b.recreador_id] = (b.recreador.nome, b.recreador.telefone or '')

    resumos: list[ResumoDiaRH] = []
    for (rid, dia), lista in sorted(por_chave.items(), key=lambda x: (x[0][1], nomes.get(x[0][0], ('', ''))[0])):
        ordenadas = sorted(lista, key=lambda x: x.registrado_em)
        entradas = [b for b in ordenadas if b.tipo == TipoPontoBatida.ENTRADA]
        saidas = [b for b in ordenadas if b.tipo == TipoPontoBatida.SAIDA]
        horas, obs = _horas_pares(ordenadas)
        nome, tel = nomes.get(rid, ('—', ''))
        resumos.append(
            ResumoDiaRH(
                recreador_id=rid,
                recreador_nome=nome,
                telefone=tel,
                data=dia,
                primeira_entrada=_fmt_hora(entradas[0].registrado_em) if entradas else '—',
                ultima_saida=_fmt_hora(saidas[-1].registrado_em) if saidas else '—',
                horas_trabalhadas=horas,
                qtd_entradas=len(entradas),
                qtd_saidas=len(saidas),
                qtd_extras=sum(1 for b in ordenadas if b.extra_plantao),
                observacao=obs,
            )
        )
    return resumos


def buscar_batidas_periodo(
    hotel: Hotel,
    data_inicio: date,
    data_fim: date,
    recreador_id: int | None = None,
) -> list[PontoBatida]:
    inicio = timezone.make_aware(datetime.combine(data_inicio, datetime.min.time()))
    fim = timezone.make_aware(datetime.combine(data_fim + timedelta(days=1), datetime.min.time()))
    qs: QuerySet = (
        PontoBatida.objects.filter(hotel=hotel, registrado_em__gte=inicio, registrado_em__lt=fim)
        .select_related('recreador', 'registrado_por')
        .order_by('registrado_em')
    )
    if recreador_id:
        qs = qs.filter(recreador_id=recreador_id)
    return list(qs)


def _safe_filename(text: str) -> str:
    cleaned = re.sub(r'[^\w\-.]+', '_', text.strip(), flags=re.UNICODE)
    return cleaned[:80] or 'ponto'


def nome_arquivo_ponto(hotel: Hotel, data_inicio: date, data_fim: date, ext: str) -> str:
    slug = _safe_filename(hotel.slug or hotel.nome)
    if data_inicio == data_fim:
        periodo = data_inicio.strftime('%Y-%m-%d')
    else:
        periodo = f'{data_inicio.strftime("%Y-%m-%d")}_a_{data_fim.strftime("%Y-%m-%d")}'
    return f'ponto_RH_{slug}_{periodo}.{ext}'


def exportar_ponto_xlsx(
    hotel: Hotel,
    data_inicio: date,
    data_fim: date,
    recreador_id: int | None = None,
) -> bytes:
    batidas = buscar_batidas_periodo(hotel, data_inicio, data_fim, recreador_id)
    resumos = montar_resumos_rh(batidas, data_inicio=data_inicio, data_fim=data_fim)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resumo RH'

    primaria = PatternFill('solid', fgColor='165A38')
    header_font = Font(bold=True, color='FFFFFF')
    titulo_font = Font(bold=True, size=14, color='165A38')

    periodo_txt = (
        data_inicio.strftime('%d/%m/%Y')
        if data_inicio == data_fim
        else f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
    )
    ws['A1'] = f'Relatório de Ponto — RH / Pagamento'
    ws['A1'].font = titulo_font
    ws.merge_cells('A1:J1')
    ws['A2'] = f'Hotel: {hotel.nome}'
    ws['A3'] = f'Período: {periodo_txt}'
    ws['A4'] = f'Gerado em: {timezone.localtime().strftime("%d/%m/%Y %H:%M")}'
    ws['A5'] = 'Use a aba Resumo RH para pagamento. A aba Detalhe lista cada batida.'

    headers = [
        'Recreador',
        'WhatsApp',
        'Data',
        '1ª Entrada',
        'Última Saída',
        'Horas trabalhadas',
        'Qtd entradas',
        'Qtd saídas',
        'Batidas extra/plantão',
        'Situação',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(8, col, h)
        cell.fill = primaria
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.border = _THIN

    for row_i, r in enumerate(resumos, 9):
        valores = [
            r.recreador_nome,
            r.telefone or '—',
            r.data.strftime('%d/%m/%Y'),
            r.primeira_entrada,
            r.ultima_saida,
            r.horas_trabalhadas,
            r.qtd_entradas,
            r.qtd_saidas,
            r.qtd_extras,
            r.observacao,
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row_i, col, val)
            cell.border = _THIN
            if col == 6:
                cell.number_format = '0.00'

    if not resumos:
        ws.cell(9, 1, 'Nenhuma batida no período.')

    widths = [28, 16, 12, 12, 12, 16, 12, 12, 18, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[8].height = 30

    # Totais
    total_row = 9 + max(len(resumos), 1)
    ws.cell(total_row + 1, 1, 'TOTAL HORAS (soma do resumo)').font = Font(bold=True)
    ws.cell(total_row + 1, 6, round(sum(r.horas_trabalhadas for r in resumos), 2)).font = Font(bold=True)
    ws.cell(total_row + 1, 6).number_format = '0.00'

    # Detalhe
    wd = wb.create_sheet('Detalhe batidas')
    wd['A1'] = 'Detalhe de batidas — conferência RH'
    wd['A1'].font = titulo_font
    wd.merge_cells('A1:G1')
    det_headers = ['Data', 'Hora', 'Recreador', 'Tipo', 'Extra/Plantão', 'IP', 'Registrado por']
    for col, h in enumerate(det_headers, 1):
        cell = wd.cell(3, col, h)
        cell.fill = primaria
        cell.font = header_font
        cell.border = _THIN
    for row_i, b in enumerate(batidas, 4):
        local = timezone.localtime(b.registrado_em)
        vals = [
            local.strftime('%d/%m/%Y'),
            local.strftime('%H:%M:%S'),
            b.recreador.nome,
            b.get_tipo_display(),
            'Sim' if b.extra_plantao else 'Não',
            b.ip or '—',
            b.registrado_por.get_username() if b.registrado_por_id else 'próprio / quiosque',
        ]
        for col, val in enumerate(vals, 1):
            cell = wd.cell(row_i, col, val)
            cell.border = _THIN
    if not batidas:
        wd.cell(4, 1, 'Nenhuma batida no período.')
    for i, w in enumerate([12, 12, 28, 10, 14, 16, 22], 1):
        wd.column_dimensions[get_column_letter(i)].width = w

    # Instruções
    wi = wb.create_sheet('Instruções RH')
    wi['A1'] = 'Como usar este arquivo'
    wi['A1'].font = titulo_font
    instrucoes = [
        '1. Aba "Resumo RH": uma linha por recreador e por dia — base para pagamento.',
        '2. Coluna "Horas trabalhadas": soma dos intervalos Entrada → Saída do dia.',
        '3. "Situação" diferente de OK indica inconsistência (ex.: saída sem entrada) — conferir na aba Detalhe.',
        '4. "Batidas extra/plantão": quantidade de batidas marcadas como extra no dia.',
        '5. Aba "Detalhe batidas": lista cronológica para auditoria.',
        '6. Filtros aplicados no sistema (hotel / período / recreador) já estão refletidos neste arquivo.',
    ]
    for i, linha in enumerate(instrucoes, 3):
        wi.cell(i, 1, linha)
    wi.column_dimensions['A'].width = 100

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_ponto_pdf(
    hotel: Hotel,
    data_inicio: date,
    data_fim: date,
    recreador_id: int | None = None,
) -> bytes:
    batidas = buscar_batidas_periodo(hotel, data_inicio, data_fim, recreador_id)
    resumos = montar_resumos_rh(batidas, data_inicio=data_inicio, data_fim=data_fim)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=f'Ponto RH — {hotel.nome}',
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        'TituloPonto',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#165A38'),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    sub = ParagraphStyle(
        'SubPonto',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    body = ParagraphStyle('BodyPonto', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT)

    periodo_txt = (
        data_inicio.strftime('%d/%m/%Y')
        if data_inicio == data_fim
        else f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
    )

    story = [
        Paragraph('Relatório de Ponto — RH / Pagamento', title),
        Paragraph(f'<b>{hotel.nome}</b> · Período: {periodo_txt}', sub),
        Paragraph(
            f'Gerado em {timezone.localtime().strftime("%d/%m/%Y %H:%M")} · '
            'Resumo por recreador/dia para pagamento. Situação ≠ OK exige conferência.',
            sub,
        ),
        Spacer(1, 0.4 * cm),
    ]

    header = [
        'Recreador',
        'WhatsApp',
        'Data',
        '1ª Entrada',
        'Última Saída',
        'Horas',
        'Ent.',
        'Saí.',
        'Extra',
        'Situação',
    ]
    data_table = [header]
    for r in resumos:
        data_table.append([
            Paragraph(r.recreador_nome[:40], body),
            r.telefone or '—',
            r.data.strftime('%d/%m/%Y'),
            r.primeira_entrada,
            r.ultima_saida,
            f'{r.horas_trabalhadas:.2f}',
            str(r.qtd_entradas),
            str(r.qtd_saidas),
            str(r.qtd_extras),
            Paragraph(r.observacao[:50], body),
        ])
    if not resumos:
        data_table.append(['Nenhuma batida no período.'] + [''] * 9)

    total_horas = round(sum(r.horas_trabalhadas for r in resumos), 2)
    data_table.append([
        Paragraph('<b>TOTAL</b>', body),
        '',
        '',
        '',
        '',
        f'{total_horas:.2f}',
        '',
        '',
        '',
        '',
    ])

    col_widths = [5.2 * cm, 2.8 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 1.8 * cm, 1.3 * cm, 1.3 * cm, 1.5 * cm, 4.5 * cm]
    table = Table(data_table, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#165A38')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (3, 1), (8, -1), 'CENTER'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F0E4')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F7FAF8')]),
        ])
    )
    story.append(table)
    story.append(Spacer(1, 0.5 * cm))
    story.append(
        Paragraph(
            f'<b>Detalhe:</b> {len(batidas)} batida(s) no período. '
            'Para a lista completa use a planilha Excel (aba Detalhe batidas).',
            body,
        )
    )

    doc.build(story)
    return buf.getvalue()
