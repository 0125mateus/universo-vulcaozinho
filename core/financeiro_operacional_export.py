"""Exportação de planilhas XLSX — financeiro operacional."""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import ExtraRecreador, ItemCompraSemanal, PagamentoAtracao, PeriodoOperacional

DIAS_SEMANA = (
    ('seg', 'Seg'),
    ('ter', 'Ter'),
    ('qua', 'Qua'),
    ('qui', 'Qui'),
    ('sex', 'Sex'),
    ('sab', 'Sáb'),
    ('dom', 'Dom'),
)

_MONEY_FMT = '#,##0.00'
_THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


@dataclass(frozen=True)
class TemaHotelPlanilha:
    primaria: str
    secundaria: str
    destaque: str
    texto_primaria: str
    texto_secundaria: str
    linha_clara: str
    linha_destaque: str


def _hex_to_xlsx(hex_color: str, fallback: str = '1E6B43') -> str:
    cleaned = re.sub(r'[^0-9A-Fa-f]', '', hex_color or '')
    if len(cleaned) == 6:
        return cleaned.upper()
    if len(cleaned) == 3:
        return ''.join(ch * 2 for ch in cleaned).upper()
    return fallback


def _luminance(rgb_hex: str) -> float:
    r = int(rgb_hex[0:2], 16) / 255
    g = int(rgb_hex[2:4], 16) / 255
    b = int(rgb_hex[4:6], 16) / 255
    return 0.2126 * r + 0.7152 * g + 0.0722 * b


def _contrast_text(bg_hex: str) -> str:
    return 'FFFFFF' if _luminance(bg_hex) < 0.55 else '1A1A1A'


def _light_tint(hex_color: str, factor: float = 0.88) -> str:
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'{r:02X}{g:02X}{b:02X}'


def _tema_do_periodo(periodo: PeriodoOperacional) -> TemaHotelPlanilha:
    hotel = periodo.hotel if periodo.hotel_id else None
    primaria = _hex_to_xlsx(hotel.cor_primaria if hotel else '#1E6B43')
    secundaria = _hex_to_xlsx(hotel.cor_secundaria if hotel else '#2FAE63', primaria)
    destaque = _hex_to_xlsx(
        getattr(hotel, 'cor_destaque', '#FFD43D') if hotel else '#FFD43D',
        secundaria,
    )
    return TemaHotelPlanilha(
        primaria=primaria,
        secundaria=secundaria,
        destaque=destaque,
        texto_primaria=_contrast_text(primaria),
        texto_secundaria=_contrast_text(secundaria),
        linha_clara=_light_tint(primaria, 0.92),
        linha_destaque=_light_tint(secundaria, 0.82),
    )


def _safe_filename(text: str) -> str:
    cleaned = re.sub(r'[^\w\-.]+', '_', text.strip(), flags=re.UNICODE)
    return cleaned[:80] or 'planilha'


def _workbook_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)


def _merge_banner(ws, row: int, col_count: int) -> None:
    if col_count > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)


def _style_banner_cell(cell, tema: TemaHotelPlanilha, bg: str, text_color: str, size: int = 14) -> None:
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, size=size, color=text_color)
    cell.alignment = Alignment(horizontal='left', vertical='center')


def _style_header_row(ws, row: int, cols: int, tema: TemaHotelPlanilha) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=tema.texto_primaria)
        cell.fill = _fill(tema.primaria)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _THIN_BORDER


def _style_data_cell(cell, tema: TemaHotelPlanilha, striped: bool = False) -> None:
    if striped:
        cell.fill = _fill(tema.linha_clara)
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical='center')


def _style_total_cell(cell, tema: TemaHotelPlanilha) -> None:
    cell.font = Font(bold=True, color=tema.texto_secundaria)
    cell.fill = _fill(tema.secundaria)
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _hotel_prefix(periodo: PeriodoOperacional) -> str:
    if periodo.hotel_id and periodo.hotel:
        return f'{_safe_filename(periodo.hotel.nome)}_'
    return ''


def _write_periodo_info(
    ws,
    periodo: PeriodoOperacional,
    titulo: str,
    tema: TemaHotelPlanilha,
    col_count: int,
) -> int:
    hotel = periodo.hotel if periodo.hotel_id else None

    _merge_banner(ws, 1, col_count)
    _style_banner_cell(ws['A1'], tema, tema.primaria, tema.texto_primaria, size=15)
    ws['A1'] = titulo
    ws.row_dimensions[1].height = 28

    row = 2
    if hotel:
        _merge_banner(ws, row, col_count)
        cell = ws[f'A{row}']
        cell.value = f'Hotel: {hotel.nome}'
        _style_banner_cell(cell, tema, tema.secundaria, tema.texto_secundaria, size=12)
        ws.row_dimensions[row].height = 22
        row += 1
        if hotel.cidade:
            _merge_banner(ws, row, col_count)
            cell = ws[f'A{row}']
            cell.value = f'Local: {hotel.cidade}/{hotel.estado}'
            cell.fill = _fill(tema.linha_destaque)
            cell.font = Font(bold=True, color=tema.texto_primaria)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1

    info_lines = [
        f'Período: {periodo.titulo}',
        (
            f'{periodo.data_inicio.strftime("%d/%m/%Y")} a '
            f'{periodo.data_fim.strftime("%d/%m/%Y")}'
        ),
    ]
    if periodo.ocupacao_pct is not None:
        info_lines.append(f'Ocupação: {periodo.ocupacao_pct}%')
    if periodo.qtd_pax:
        info_lines.append(f'Hóspedes: {periodo.qtd_pax}')

    for line in info_lines:
        _merge_banner(ws, row, col_count)
        cell = ws[f'A{row}']
        cell.value = line
        cell.fill = _fill(tema.linha_clara)
        cell.font = Font(color='333333')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    return row + 1


def exportar_extras_recreadores_xlsx(periodo: PeriodoOperacional) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Extras recreadores'
    tema = _tema_do_periodo(periodo)
    col_count = 9

    header_row = _write_periodo_info(ws, periodo, 'Extras de recreadores', tema, col_count)
    headers = ['Recreador', *[label for _, label in DIAS_SEMANA], 'Total']
    for col, title in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=title)
    _style_header_row(ws, header_row, len(headers), tema)

    totais_dia = {dia: Decimal('0') for dia, _ in DIAS_SEMANA}
    data_row = header_row + 1
    extras = periodo.extras_recreadores.all()

    for i, extra in enumerate(extras):
        striped = i % 2 == 1
        nome_cell = ws.cell(row=data_row, column=1, value=extra.nome)
        _style_data_cell(nome_cell, tema, striped=striped)
        for col, (dia, _) in enumerate(DIAS_SEMANA, start=2):
            valor = getattr(extra, f'valor_{dia}')
            cell = ws.cell(row=data_row, column=col, value=float(valor))
            cell.number_format = _MONEY_FMT
            _style_data_cell(cell, tema, striped=striped)
            totais_dia[dia] += valor
        total_cell = ws.cell(row=data_row, column=9, value=float(extra.total))
        total_cell.number_format = _MONEY_FMT
        _style_data_cell(total_cell, tema, striped=striped)
        total_cell.font = Font(bold=True)
        data_row += 1

    footer_row = data_row
    total_label = ws.cell(row=footer_row, column=1, value='Totais do dia')
    _style_total_cell(total_label, tema)
    total_label.alignment = Alignment(horizontal='left', vertical='center')
    for col, (dia, _) in enumerate(DIAS_SEMANA, start=2):
        cell = ws.cell(row=footer_row, column=col, value=float(totais_dia[dia]))
        cell.number_format = _MONEY_FMT
        _style_total_cell(cell, tema)
    total_geral = sum(totais_dia.values())
    total_cell = ws.cell(row=footer_row, column=9, value=float(total_geral))
    total_cell.number_format = _MONEY_FMT
    _style_total_cell(total_cell, tema)

    ws.column_dimensions['A'].width = 22
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['I'].width = 14

    return _workbook_bytes(wb)


def exportar_atracoes_xlsx(periodo: PeriodoOperacional) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Atrações'
    tema = _tema_do_periodo(periodo)
    col_count = 9

    header_row = _write_periodo_info(ws, periodo, 'Pagamentos de atrações / artistas', tema, col_count)
    headers = [
        'Data', 'Artista', 'Atração', 'Valor (R$)', 'Chave PIX',
        'Evento', 'Status', 'Autorização', 'Observações',
    ]
    for col, title in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=title)
    _style_header_row(ws, header_row, len(headers), tema)

    data_row = header_row + 1
    total = Decimal('0')
    pagamentos = PagamentoAtracao.objects.filter(periodo=periodo).order_by('data_evento', 'artista')

    for i, pag in enumerate(pagamentos):
        striped = i % 2 == 1
        data_txt = ''
        if pag.data_evento:
            data_txt = pag.data_evento.strftime('%d/%m/%Y')
        elif pag.data_label:
            data_txt = pag.data_label

        values = [
            data_txt, pag.artista, pag.atracao, float(pag.valor), pag.chave_pix,
            pag.evento, pag.get_status_display(), pag.autorizacao_diretoria, pag.observacoes,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col, value=value)
            _style_data_cell(cell, tema, striped=striped)
            if col == 4:
                cell.number_format = _MONEY_FMT
        total += pag.valor
        data_row += 1

    total_label = ws.cell(row=data_row, column=3, value='TOTAL')
    _style_total_cell(total_label, tema)
    total_cell = ws.cell(row=data_row, column=4, value=float(total))
    total_cell.number_format = _MONEY_FMT
    _style_total_cell(total_cell, tema)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 30

    return _workbook_bytes(wb)


def exportar_compras_xlsx(periodo: PeriodoOperacional) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compras'
    tema = _tema_do_periodo(periodo)
    col_count = 5

    header_row = _write_periodo_info(ws, periodo, 'Compras semanais de materiais', tema, col_count)
    headers = ['Material', 'Quantidade', 'Link fornecedor', 'Preço unitário', 'Preço total']
    for col, title in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=title)
    _style_header_row(ws, header_row, len(headers), tema)

    data_row = header_row + 1
    total = Decimal('0')
    itens = ItemCompraSemanal.objects.filter(periodo=periodo).order_by('ordem', 'descricao')

    for i, item in enumerate(itens):
        striped = i % 2 == 1
        values = [
            item.descricao,
            item.quantidade,
            item.link_fornecedor,
            float(item.preco_unitario),
            float(item.preco_total),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col, value=value)
            _style_data_cell(cell, tema, striped=striped)
            if col in (4, 5):
                cell.number_format = _MONEY_FMT
        total += item.preco_total
        data_row += 1

    total_label = ws.cell(row=data_row, column=4, value='TOTAL')
    _style_total_cell(total_label, tema)
    total_cell = ws.cell(row=data_row, column=5, value=float(total))
    total_cell.number_format = _MONEY_FMT
    _style_total_cell(total_cell, tema)

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    return _workbook_bytes(wb)


def nome_arquivo_extras(periodo: PeriodoOperacional) -> str:
    return f'{_hotel_prefix(periodo)}{_safe_filename(periodo.titulo)}_extras_recreadores.xlsx'


def nome_arquivo_atracoes(periodo: PeriodoOperacional) -> str:
    return f'{_hotel_prefix(periodo)}{_safe_filename(periodo.titulo)}_atracoes.xlsx'


def nome_arquivo_compras(periodo: PeriodoOperacional) -> str:
    return f'{_hotel_prefix(periodo)}{_safe_filename(periodo.titulo)}_compras.xlsx'
